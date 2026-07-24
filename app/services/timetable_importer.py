from __future__ import annotations

import base64
import csv
import io
import json
import re
from typing import Any

from openpyxl import load_workbook
from pypdf import PdfReader

from app.schemas.timetable import CourseSessionCreate


HEADER_ALIASES = {
    "course_name": ("课程名称", "课程", "course_name", "course"),
    "weekday": ("星期", "周几", "weekday", "day"),
    "start_period": ("开始节次", "起始节次", "start_period", "start"),
    "end_period": ("结束节次", "终止节次", "end_period", "end"),
    "location": ("地点", "教室", "location", "room"),
    "weeks": ("周次", "教学周", "weeks", "week"),
}

WEEKDAY_VALUES = {
    "一": 1,
    "二": 2,
    "三": 3,
    "四": 4,
    "五": 5,
    "六": 6,
    "日": 7,
    "天": 7,
    "周一": 1,
    "周二": 2,
    "周三": 3,
    "周四": 4,
    "周五": 5,
    "周六": 6,
    "周日": 7,
    "星期一": 1,
    "星期二": 2,
    "星期三": 3,
    "星期四": 4,
    "星期五": 5,
    "星期六": 6,
    "星期日": 7,
    "星期天": 7,
}

# 杭电教务系统导出的课表 PDF 使用固定七列。这里仅把列坐标用于识别
# “星期几”，课程名称、节次、周次和地点仍从 PDF 文本中读取。
HDU_WEEKDAY_COLUMN_X = {
    1: 104.1,
    2: 207.9,
    3: 311.8,
    4: 415.6,
    5: 519.5,
    6: 623.3,
    7: 727.1,
}
HDU_COLUMN_TOLERANCE = 18.0


def parse_timetable(
    *,
    content: str,
    format_name: str,
) -> tuple[list[CourseSessionCreate], int, list[str]]:
    rows = _load_rows(content=content, format_name=format_name)
    entries: list[CourseSessionCreate] = []
    skipped = 0
    messages: list[str] = []
    for index, raw in enumerate(rows, start=2):
        try:
            normalized = _normalize_row(raw)
            if not normalized["course_name"]:
                skipped += 1
                continue
            entries.append(CourseSessionCreate.model_validate(normalized))
        except Exception as exc:
            skipped += 1
            messages.append(f"第{index}行未导入：{exc}")
    if not entries:
        detail = messages[0] if messages else "文件中没有可识别的课程"
        raise ValueError(detail)
    return entries, skipped, messages[:10]


def _load_rows(*, content: str, format_name: str) -> list[dict[str, Any]]:
    if format_name == "json":
        payload = json.loads(content)
        rows = payload.get("entries", []) if isinstance(payload, dict) else payload
        if not isinstance(rows, list):
            raise ValueError("JSON需要是课程数组，或包含entries数组")
        return [row for row in rows if isinstance(row, dict)]
    if format_name == "csv":
        return list(csv.DictReader(io.StringIO(content.lstrip("\ufeff"))))
    if format_name == "xlsx_base64":
        binary = base64.b64decode(content, validate=True)
        workbook = load_workbook(io.BytesIO(binary), read_only=True, data_only=True)
        sheet = workbook.active
        values = list(sheet.iter_rows(values_only=True))
        if not values:
            return []
        headers = [str(value or "").strip() for value in values[0]]
        return [
            {
                headers[index]: value
                for index, value in enumerate(row)
                if index < len(headers) and headers[index]
            }
            for row in values[1:]
        ]
    if format_name == "pdf_base64":
        binary = base64.b64decode(content, validate=True)
        if len(binary) > 5_000_000:
            raise ValueError("课表PDF不能超过5MB")
        return _load_hdu_pdf_rows(binary)
    raise ValueError("暂不支持这种课表格式")


def _load_hdu_pdf_rows(binary: bytes) -> list[dict[str, Any]]:
    try:
        reader = PdfReader(io.BytesIO(binary))
    except Exception as exc:
        raise ValueError("PDF文件无法读取或已经损坏") from exc
    if len(reader.pages) > 20:
        raise ValueError("课表PDF页数过多，请只上传当前学期的课表")
    positioned_pages: list[list[tuple[float, float, float, str]]] = []
    for page in reader.pages:
        positioned: list[tuple[float, float, float, str]] = []

        def visitor_text(
            text,
            cm,
            tm,
            font_dict,
            font_size,
        ) -> None:
            compact = " ".join(str(text or "").split())
            if not compact:
                return
            positioned.append(
                (
                    float(tm[4]),
                    float(tm[5]),
                    float(font_size),
                    compact,
                )
            )

        page.extract_text(visitor_text=visitor_text)
        positioned_pages.append(positioned)
    rows = _rows_from_hdu_positioned_pages(positioned_pages)
    if not rows:
        raise ValueError(
            "没有识别到杭电教务系统课表网格；其他学校请先使用"
            "Excel/CSV/JSON通用模板"
        )
    return rows


def _rows_from_hdu_positioned_pages(
    pages: list[list[tuple[float, float, float, str]]],
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for positioned in pages:
        current: dict[str, Any] | None = None
        for x, _y, font_size, text in positioned:
            weekday = _weekday_from_column_x(x)
            is_course_title = (
                weekday is not None
                and 8.5 <= font_size <= 9.5
                and not text.startswith(("(", "（"))
                and "节)" not in text
            )
            if is_course_title:
                _append_hdu_pdf_row(rows, current)
                current = {
                    "course_name": text,
                    "weekday": weekday,
                    "column_x": x,
                    "details": [],
                }
                continue
            if (
                current is not None
                and abs(x - float(current["column_x"])) <= 3.0
                and font_size < 8.5
            ):
                current["details"].append(text)
        _append_hdu_pdf_row(rows, current)
    return _merge_duplicate_pdf_rows(rows)


def _weekday_from_column_x(value: float) -> int | None:
    weekday, column_x = min(
        HDU_WEEKDAY_COLUMN_X.items(),
        key=lambda item: abs(item[1] - value),
    )
    return (
        weekday
        if abs(column_x - value) <= HDU_COLUMN_TOLERANCE
        else None
    )


def _append_hdu_pdf_row(
    rows: list[dict[str, Any]],
    current: dict[str, Any] | None,
) -> None:
    if current is None:
        return
    details = "".join(current["details"])
    period_match = re.search(
        r"[（(]\s*(\d{1,2})\s*[-—~至]\s*(\d{1,2})\s*节\s*[）)]"
        r"(.*?)(?=/校区:|/场地:|/教师:|$)",
        details,
    )
    if not period_match:
        return
    location_match = re.search(
        r"(?:^|/)场地:(.*?)(?=/教师:|/教学班:|$)",
        details,
    )
    location = (
        location_match.group(1).strip()
        if location_match
        else None
    )
    if location and "不在教室" in location:
        location = None
    rows.append(
        {
            "course_name": current["course_name"],
            "weekday": current["weekday"],
            "start_period": int(period_match.group(1)),
            "end_period": int(period_match.group(2)),
            "location": location,
            "weeks": period_match.group(3),
        }
    )


def _merge_duplicate_pdf_rows(
    rows: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    merged: dict[tuple[Any, ...], dict[str, Any]] = {}
    for row in rows:
        key = (
            row["course_name"],
            row["weekday"],
            row["start_period"],
            row["end_period"],
            row.get("location"),
        )
        if key not in merged:
            merged[key] = {
                **row,
                "weeks": _weeks_value(row.get("weeks")),
            }
            continue
        merged[key]["weeks"] = sorted(
            set(merged[key]["weeks"])
            | set(_weeks_value(row.get("weeks")))
        )
    return list(merged.values())


def _normalize_row(raw: dict[str, Any]) -> dict[str, Any]:
    resolved = {
        key: _first_value(raw, aliases)
        for key, aliases in HEADER_ALIASES.items()
    }
    weekday_raw = str(resolved["weekday"] or "").strip()
    weekday = WEEKDAY_VALUES.get(weekday_raw)
    if weekday is None and weekday_raw.isdigit():
        weekday = int(weekday_raw)
    start_period = _period_value(resolved["start_period"])
    end_period = _period_value(resolved["end_period"]) or start_period
    return {
        "course_name": str(resolved["course_name"] or "").strip(),
        "weekday": weekday,
        "start_period": start_period,
        "end_period": end_period,
        "location": str(resolved["location"] or "").strip() or None,
        "weeks": _weeks_value(resolved["weeks"]),
    }


def _first_value(raw: dict[str, Any], aliases: tuple[str, ...]) -> Any:
    compact = {str(key).strip().lower(): value for key, value in raw.items()}
    for alias in aliases:
        if alias.lower() in compact:
            return compact[alias.lower()]
    return None


def _period_value(value: Any) -> int | None:
    match = re.search(r"\d{1,2}", str(value or ""))
    return int(match.group()) if match else None


def _weeks_value(value: Any) -> list[int]:
    if isinstance(value, (list, tuple, set)):
        return sorted(
            {
                int(week)
                for week in value
                if str(week).strip().isdigit()
            }
        )
    text = str(value or "").strip()
    if not text:
        return []
    weeks: set[int] = set()
    for segment in re.split(r"[,，、;；]", text):
        for start_raw, end_raw, parity in re.findall(
            r"(\d{1,2})(?:\s*[-—~至]\s*(\d{1,2}))?\s*周?"
            r"(?:\s*[（(]\s*([单双])\s*[）)])?",
            segment,
        ):
            start = int(start_raw)
            end = int(end_raw or start_raw)
            candidates = range(min(start, end), max(start, end) + 1)
            if parity == "单":
                candidates = (week for week in candidates if week % 2 == 1)
            elif parity == "双":
                candidates = (week for week in candidates if week % 2 == 0)
            weeks.update(candidates)
    return sorted(weeks)
